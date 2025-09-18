#main libraries
import customtkinter as ctk
import jdatetime
from datetime import datetime
from datetime import timedelta
from tkinter.filedialog import asksaveasfilename
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

#data imports
import asset_paths
from bol_fake_data import current_bols
from insystem_data import current_deals, old_deals
from bol_inf import BolInf
import scrollbar

#controllers
import common_controller as cc
import common_ctk as ck
import register_controller as rc

#colors
import colors

class BolBox(ctk.CTkFrame):
    def __init__(self, parent, box, bol, width, y):
        self.parent = parent
        self.box = box 
        self.bol = bol
        self.width = width
        self.y = y
        self.height = 60
           
        main_box = ck.make_frame(box, self.width, self.height, colors.dark_green_6, colors.green_5, colors.green_5,
                                0, 20, 0.05, y, "nw")
        main_box.pack(fill="x", padx=5, pady=5)
        name_label = ck.make_label(main_box, 50, 30, colors.green_5, colors.green_5, colors.black, "شماره بارنامه: " + cc.eng_to_persian_date(self.bol["load_id"]), 0, None, 
                    (None, 15, "bold"), 0.97, 0.3, "e")
        des_1_label = ck.make_label(main_box, 50, 30, colors.green_5, colors.green_5, colors.black, "نام راننده: " + self.bol["driver_name"], 0, None, 
                    (None, 12, "bold"), 0.53, 0.3, "e")
        des_2_label = ck.make_label(main_box, 50, 20, colors.green_5, colors.green_5, colors.black, "نام فرآورده: " + bol["load_name"], 0, "right", 
                    (None, 10, "bold"), 0.5, 0.65, "center")
        def hover_box():
            main_box.configure(fg_color = colors.green_4)
            name_label.configure(fg_color = colors.green_4)
            des_1_label.configure(fg_color = colors.green_4)
            des_2_label.configure(fg_color = colors.green_4)
        def unhover_box():
            main_box.configure(fg_color = colors.green_5)
            name_label.configure(fg_color = colors.green_5)
            des_1_label.configure(fg_color = colors.green_5)
            des_2_label.configure(fg_color = colors.green_5)
          
        def on_deal_clicked(event):
            global bol_frame
            try: bol_frame.destroy()
            except: pass
            bol_frame = ctk.CTkFrame(self.parent, 
                                     bg_color=colors.light_green_1,
                                     fg_color=colors.light_green_1,
                                     width=1050, 
                                     height=400)
            bol_frame.place(relx=0.5, rely=0.5, anchor="center")

            cal = BolInf(bol_frame, None, self.bol)
            cal.place(relx=0.5, rely=0.5, anchor="center")
            
            def do_delete(): 
                bol_frame.destroy()
            close_btn = ctk.CTkButton(bol_frame, text="بستن", 
                                        width=120,
                                        height=30, 
                                        font=(None, 25),
                                        bg_color=colors.light_green_1, 
                                        corner_radius=60,
                                    fg_color=colors.dark_green_6,
                                    text_color=colors.light_green_1,
                                    hover_color=colors.green_3,
                                    command=do_delete)
            close_btn.place(relx=0.075, rely=0.91, anchor="center")
              
        main_box.bind("<Enter>", lambda e: hover_box())
        name_label.bind("<Enter>", lambda e: hover_box())
        des_1_label.bind("<Enter>", lambda e: hover_box())
        des_2_label.bind("<Enter>", lambda e: hover_box())
        
        main_box.bind("<Leave>", lambda e: unhover_box())
        name_label.bind("<Leave>", lambda e: unhover_box())
        des_1_label.bind("<Leave>", lambda e: unhover_box())
        des_2_label.bind("<Leave>", lambda e: unhover_box())
        
        main_box.bind("<Button-1>", on_deal_clicked)
        name_label.bind("<Button-1>", on_deal_clicked)
        des_1_label.bind("<Button-1>", on_deal_clicked)
        des_2_label.bind("<Button-1>", on_deal_clicked)
        
class GroupBox(ctk.CTkFrame):
    def __init__(self, parent, box, groups_box, group, group_id, width, y):
        self.parent = parent
        self.box = box 
        self.groups_box = groups_box
        self.group = group
        self.group_id = group_id
        self.width = width
        self.y = y
        self.height = 60
           
        def print_group():
            wb = Workbook()
            ws = wb.active
            columns = ['ردیف', 'تاریخ ثبت', 'قرارداد', 'شماره حواله', 'شماره بارنامه', 'شماره صورت بارنامه', 'نماینده', 'شماره صورت نماینده', 'تاریخ صدور', 'تاریخ ارسال', 'نام راننده', 'کد ملی راننده', 'شماره هوشمند راننده', 'نام مالک', 'شماره نفتکش', 'شماره هوشمند نفتکش', 'مبدأ', 'مقصد', 'وزن واقعی', 'وزن حمل شده', 'عنوان کالا', 'نام فرآورده', 'نوع ارسال', 'مقدار دریافتی', 'مقدار پرداختی']
            ws.append(columns)
            fill = PatternFill(start_color=colors.dark_green_6[1:], end_color=colors.dark_green_6[1:], fill_type="solid")
            font = Font(color=colors.white[1:], size=15,  bold=True)
            alignment = Alignment(horizontal="center", vertical="center")
            thin = Side(border_style="thin", color="000000")
            border = Border(top=thin, left=thin, right=thin, bottom=thin)
            
            total_bols = [cc.get_bol_by_load_id(bol) for bol in self.group['bols']]
            
            true_value_sum = 0
            load_weight_sum = 0
            input_money_sum = 0
            payed_money_sum = 0
            
            for i, item in enumerate(total_bols):
                new_list = list(item.values())
                true_value_sum += int(cc.persian_to_eng_date(new_list[17]))
                load_weight_sum += int(cc.persian_to_eng_date(new_list[18]))
                input_money_sum += int(cc.persian_to_eng_date(new_list[22]))
                payed_money_sum += int(cc.persian_to_eng_date(new_list[23]))
                new_list[17] = cc.cash_format(new_list[17])
                new_list[18] = cc.cash_format(new_list[18])
                new_list[22] = cc.cash_format(new_list[22])
                new_list[23] = cc.cash_format(new_list[23])
                tpl = (i + 1, ) + tuple(new_list)
                new_tpl = tuple(cc.eng_to_persian_date(str(x)) for x in tpl)
                ws.append(new_tpl)

            ws.cell(row=len(total_bols) + 2, column=19, value=cc.eng_to_persian_date(cc.cash_format(str(true_value_sum))))
            ws.cell(row=len(total_bols) + 2, column=20, value=cc.eng_to_persian_date(cc.cash_format(str(load_weight_sum))))
            ws.cell(row=len(total_bols) + 2, column=24, value=cc.eng_to_persian_date(cc.cash_format(str(input_money_sum))))
            ws.cell(row=len(total_bols) + 2, column=25, value=cc.eng_to_persian_date(cc.cash_format(str(payed_money_sum))))

            for col in range(1, 26):
                col_letter = ws.cell(row=1, column=col).column_letter
                if col == 1: ws.column_dimensions[col_letter].width = 8
                if col in [19, 20, 21, 22, 23, 25, 26]: ws.column_dimensions[col_letter].width = 15
                if col in [2, 3, 4, 5, 6, 8, 9, 10, 11, 12, 13, 14, 15, 16, 24]: ws.column_dimensions[col_letter].width = 20
                if col in [7, 17, 18]: ws.column_dimensions[col_letter].width = 30
            #مشخص کردن رنگ و فونت هدینگ
            for col_num in range(1, len(columns)+1):
                ws.cell(row=1, column=col_num).fill = fill
                ws.cell(row=1, column=col_num).font = font
            for row_id in range(1 + len(total_bols)):
                ws.cell(row=row_id+1, column=1).fill = fill
                ws.cell(row=row_id+1, column=1).font = font
            for col_id in range(len(columns)):
                for row_id in range(1 + len(total_bols)):
                    ws.cell(row=row_id+1, column=col_id+1).border = border
                    ws.cell(row=row_id+1, column=col_id+1).alignment = alignment
            
            ws.cell(row=2 + len(total_bols), column=19).fill = fill
            ws.cell(row=2 + len(total_bols), column=19).font = font
            ws.cell(row=2 + len(total_bols), column=20).fill = fill
            ws.cell(row=2 + len(total_bols), column=20).font = font
            ws.cell(row=2 + len(total_bols), column=24).fill = fill
            ws.cell(row=2 + len(total_bols), column=24).font = font
            ws.cell(row=2 + len(total_bols), column=25).fill = fill
            ws.cell(row=2 + len(total_bols), column=25).font = font
            
            ws.cell(row=2 + len(total_bols), column=19).border = border
            ws.cell(row=2 + len(total_bols), column=19).alignment = alignment
            ws.cell(row=2 + len(total_bols), column=20).border = border
            ws.cell(row=2 + len(total_bols), column=20).alignment = alignment
            ws.cell(row=2 + len(total_bols), column=24).border = border
            ws.cell(row=2 + len(total_bols), column=24).alignment = alignment
            ws.cell(row=2 + len(total_bols), column=25).border = border
            ws.cell(row=2 + len(total_bols), column=25).alignment = alignment
            
            
            
            file_path = asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")]
            )
            if file_path:
                wb.save(file_path)
            
        self.main_box = ck.make_frame(box, self.width, self.height, colors.green_5, colors.dark_green_6, colors.dark_green_6,
                                0, 20, 0.05, y, "nw")
        self.main_box.pack(fill="x", padx=5, pady=5)
        
        name_label = ck.make_label(self.main_box, 50, 30, colors.dark_green_6, colors.dark_green_6, colors.white, "صورت شماره " + cc.eng_to_persian_date(group['id']), 0, None, 
                    (None, 15, "bold"), 0.97, 0.3, "e")
        des_1_label = ck.make_label(self.main_box, 50, 30, colors.dark_green_6, colors.dark_green_6, colors.white, "تعداد بارنامه های ثبت شده: " + cc.eng_to_persian_date(str(len(self.group["bols"]))), 0, None, 
                    (None, 12, "bold"), 0.53, 0.3, "e")
        des_2_label = ck.make_label(self.main_box, 50, 20, colors.dark_green_6, colors.dark_green_6, colors.white, "تاریخ ثبت صورت: " + cc.eng_to_persian_date(group["submit_date"]), 0, "right", 
                    (None, 10, "bold"), 0.5, 0.65, "center")
        if(group['payed']):
            payed_label = ck.make_label(self.main_box, 50, 20, colors.dark_green_6, colors.dark_green_6, colors.green_5, "وصول شده", 0, "right", 
                        (None, 10), 0.95, 0.65, "e")
        else:
            payed_label = ck.make_label(self.main_box, 50, 20, colors.dark_green_6, colors.dark_green_6, colors.red_color, "عدم وصول", 0, "right", 
                        (None, 10), 0.95, 0.65, "e")
        
        
        ck.make_button(self.main_box, 'چاپ', 50, 15, (None, 13, 'bold'), 20, 
                                       colors.black, colors.dark_green_6, colors.green_5, colors.green_2, 
                                       print_group, 0.2, 0.72, 'e')
        def hover_box():
            self.main_box.configure(fg_color = colors.dark_green_2)
            name_label.configure(fg_color = colors.dark_green_2)
            des_1_label.configure(fg_color = colors.dark_green_2)
            des_2_label.configure(fg_color = colors.dark_green_2)
            payed_label.configure(fg_color = colors.dark_green_2)
        def unhover_box():
            self.main_box.configure(fg_color = colors.dark_green_6)
            name_label.configure(fg_color = colors.dark_green_6)
            des_1_label.configure(fg_color = colors.dark_green_6)
            des_2_label.configure(fg_color = colors.dark_green_6)
            payed_label.configure(fg_color = colors.dark_green_6)
        
        def double_clicked(event):
            on_deal_clicked(event)
            change_group_frame = ck.make_frame(parent, 250, 180, colors.dark_green_4, colors.dark_green_4, colors.dark_green_4, 0, 0, 0.5, 0.5, 'center')
            ck.make_label(change_group_frame, 200, 40, colors.dark_green_4, colors.dark_green_4, colors.white,
                          'تغییر شماره صورت', 0, None, (None, 20, 'bold'), 0.5, 0.2, 'center')
            entry = ck.make_entry(change_group_frame, 200, 40, colors.dark_green_4, colors.green_1, colors.green_1, colors.black,
                          'شماره صورت', colors.light_gray_5, 20, "right", (None, 15, 'bold'), 0.5, 0.5, 'center')
            entry.delete(0, 'end')
            entry.insert(0, cc.eng_to_persian_date(str(group['id'])))
            
            def change_group_id():
                if(len(entry.get()) != 0): 
                    group['id'] = cc.eng_to_persian_date(entry.get())
                cancel()
                name_label.configure(text = "صورت شماره " + cc.eng_to_persian_date(group['id']))
                
            def cancel():
                change_group_frame.destroy()
            
            ck.make_button(change_group_frame, 'تأیید', 60, 30, (None, 15, 'bold'), 20, colors.black, colors.dark_green_4, colors.green_3, 
                           colors.light_green_5, change_group_id, 0.8, 0.8, 'e')
            ck.make_button(change_group_frame, 'بستن', 60, 30, (None, 15, 'bold'), 20, colors.black, colors.dark_green_4, colors.green_3, 
                           colors.light_green_5, cancel, 0.2, 0.8, 'w')
            
            
        def on_deal_clicked(event):
            for child in groups_box.winfo_children()[1:]: child.destroy()
            
            scrollable = scrollbar.ScrollableFrame(groups_box, self.width + 19, colors.dark_green_6)
            scrollable.pack(side="top", fill="x", padx=0, pady=(110, 0))
                
            for i, bol_id in enumerate(group["bols"]):
                BolBox(self.parent, scrollable.inner_frame, cc.get_bol_by_load_id(bol_id), 325, 0.2 + i * 0.17)
            return

        def on_payed_label_clicked(event):
            def do_change():
                on_deal_clicked(event)
                if(payed_label.cget('text') == 'وصول شده'):
                    group['payed'] = False
                    payed_label.configure(text = 'عدم وصول', text_color=colors.red_color)
                else: 
                    group['payed'] = True
                    payed_label.configure(text = 'وصول شده', text_color=colors.green_5)
                cancel()
            def cancel():
                frame.destroy()
            frame = rc.show_confirmation_box(parent, do_change, cancel, 'آيا از تغییر وضعیت وصول، اطمینان دارید؟', None)
                
            
        self.main_box.bind("<Enter>", lambda e: hover_box())
        name_label.bind("<Enter>", lambda e: hover_box())
        des_1_label.bind("<Enter>", lambda e: hover_box())
        des_2_label.bind("<Enter>", lambda e: hover_box())
        payed_label.bind("<Enter>", lambda e: hover_box())
        
        self.main_box.bind("<Leave>", lambda e: unhover_box())
        name_label.bind("<Leave>", lambda e: unhover_box())
        des_1_label.bind("<Leave>", lambda e: unhover_box())
        des_2_label.bind("<Leave>", lambda e: unhover_box())
        payed_label.bind("<Leave>", lambda e: unhover_box())
        
        self.main_box.bind("<Button-1>", on_deal_clicked)
        name_label.bind("<Button-1>", on_deal_clicked)
        des_1_label.bind("<Button-1>", on_deal_clicked)
        des_2_label.bind("<Button-1>", on_deal_clicked)
        payed_label.bind("<Button-1>", on_payed_label_clicked)
        
        self.main_box.bind("<Double-1>", double_clicked)
        name_label.bind("<Double-1>", double_clicked)
        des_1_label.bind("<Double-1>", double_clicked)
        des_2_label.bind("<Double-1>", double_clicked)
        payed_label.bind("<Double-1>", double_clicked)
        

class DealBox(ctk.CTkFrame):
    def __init__(self, parent, box, groups_box, bols_box, deal, width, y):
        self.parent = parent
        self.box = box 
        self.groups_box = groups_box
        self.deal = deal
        self.width = width
        self.y = y
        self.height = 60
        
        main_box = ck.make_frame(box, self.width, self.height, colors.dark_green_6, colors.green_5, colors.green_5,
                                0, 20, 0.05, y, "nw")
        main_box.pack(fill="x", padx=5, pady=5)
        name_label = ck.make_label(main_box, 50, 30, colors.green_5, colors.green_5, colors.black, self.deal["name"] + '/' + cc.eng_to_persian_date(self.deal["id"]), 0, None, 
                    (None, 15, "bold"), 0.97, 0.3, "e")
        des_1_label = ck.make_label(main_box, 50, 30, colors.green_5, colors.green_5, colors.black, "صورت های بسته شده: " + cc.eng_to_persian_date(str(len(self.deal["packages"]))), 0, None, 
                    (None, 12, "bold"), 0.45, 0.3, "e")
        today = jdatetime.date.today()

        split = list(map(int, self.deal["end_date"].split('/')))
        input_date = jdatetime.date(split[0], split[1], split[2])
        
        diff = input_date - today
        
        if(diff.days >= 0): 
            des_2_label = ck.make_label(main_box, 50, 20, colors.green_5, colors.green_5, colors.black, "مدت‌زمان باقی‌مانده از اعتبار: " + cc.eng_to_persian_date(str(diff.days)) + " روز" , 0, "right", 
                        (None, 10, "bold"), 0.5, 0.65, "center")
        else:
            des_2_label = ck.make_label(main_box, 50, 20, colors.green_5, colors.green_5, colors.red_color, "مدت‌زمان باقی‌مانده از اعتبار: " + cc.eng_to_persian_date(str(-diff.days)) + " روز از اعتبار گذشته است!" , 0, "right", 
                        (None, 10, "bold"), 0.5, 0.65, "center")
        def hover_box():
            main_box.configure(fg_color = colors.green_4)
            name_label.configure(fg_color = colors.green_4)
            des_1_label.configure(fg_color = colors.green_4)
            des_2_label.configure(fg_color = colors.green_4)
        def unhover_box():
            main_box.configure(fg_color = colors.green_5)
            name_label.configure(fg_color = colors.green_5)
            des_1_label.configure(fg_color = colors.green_5)
            des_2_label.configure(fg_color = colors.green_5)
          
        def on_deal_clicked(event):
            for child in groups_box.winfo_children()[1:]: child.destroy()
            for child in bols_box.winfo_children()[1:]: child.destroy()
            if(len(self.deal["packages"]) == 0):
                ck.make_label(groups_box, 150, 40, colors.green_5, colors.green_5, colors.black,
                            "صورت‌ای برای نمایش وجود ندارد!", 0, None, (None, 25), 0.5, 0.5, "center")
                ck.make_label(bols_box, 150, 40, colors.dark_green_6, colors.dark_green_6, colors.white,
                            "بارنامه‌ای برای نمایش وجود ندارد!", 0, None, (None, 25), 0.5, 0.5, "center")
            else:
                scrollable = scrollbar.ScrollableFrame(groups_box, self.width + 17, colors.green_5)
                scrollable.pack(side="top", fill="x", padx=0, pady=(110, 0))
                    
                for i, group in enumerate(self.deal["packages"]):
                    GroupBox(self.parent, scrollable.inner_frame, bols_box, group, i + 1, 325, 0.2 + i * 0.17)
            return
              
        main_box.bind("<Enter>", lambda e: hover_box())
        name_label.bind("<Enter>", lambda e: hover_box())
        des_1_label.bind("<Enter>", lambda e: hover_box())
        des_2_label.bind("<Enter>", lambda e: hover_box())
        
        main_box.bind("<Leave>", lambda e: unhover_box())
        name_label.bind("<Leave>", lambda e: unhover_box())
        des_1_label.bind("<Leave>", lambda e: unhover_box())
        des_2_label.bind("<Leave>", lambda e: unhover_box())
        
        main_box.bind("<Button-1>", on_deal_clicked)
        name_label.bind("<Button-1>", on_deal_clicked)
        des_1_label.bind("<Button-1>", on_deal_clicked)
        des_2_label.bind("<Button-1>", on_deal_clicked)
        

class BolGroups(ctk.CTkFrame):
    def __init__(self, parent, go_to_view_bol):
        super().__init__(parent)
        
        self.go_to_view_bol = go_to_view_bol
        self.pack(fill="both", expand=True)
        self.configure(fg_color=colors.light_green_1)
        
        main_box = ck.make_frame(self, float(self.master.winfo_screenwidth() / 1.2), 
                                 float(self.master.winfo_screenheight() / 1.5), 
                                 colors.white, colors.white, colors.white, 0, 0, 0.5, 0.5, "center")
        ck.make_label(main_box, 150, 20, colors.white, colors.white, colors.black, "صورت بارنامه ها", 0, None, 
                      (None, 45, "bold"), 0.5, 0.1, "center")
        
        right_section_box = ck.make_frame(main_box, float(self.master.winfo_screenwidth() / 4), 
                                 float(self.master.winfo_screenheight() / 2.4), 
                                 colors.dark_green_6, colors.dark_green_6, colors.dark_green_6, 0, 0, 0.97, 0.87, "se")
        middle_section_box = ck.make_frame(main_box, float(self.master.winfo_screenwidth() / 4), 
                                 float(self.master.winfo_screenheight() / 2.4), 
                                 colors.green_5, colors.green_5, colors.green_5, 0, 0, 0.65, 0.87, "se")
        left_section_box = ck.make_frame(main_box, float(self.master.winfo_screenwidth() / 4), 
                                 float(self.master.winfo_screenheight() / 2.4), 
                                 colors.dark_green_6, colors.dark_green_6, colors.dark_green_6, 0, 0, 0.33, 0.87, "se")
        ck.make_label(main_box, 150, 20, colors.white, colors.white, colors.black, "صورت بارنامه ها", 0, None, 
                      (None, 45, "bold"), 0.5, 0.1, "center")
        ck.make_label(right_section_box, 150, 20, colors.dark_green_6, colors.dark_green_6, colors.white, "قرارداد های ثبت شده", 0, None, 
                      (None, 35, "bold"), 0.5, 0.1, "center")
        ck.make_label(middle_section_box, 150, 20, colors.green_5, colors.green_5, colors.black, "صورت های بسته شده", 0, None, 
                      (None, 35, "bold"), 0.5, 0.1, "center")
        ck.make_label(left_section_box, 150, 20, colors.dark_green_6, colors.dark_green_6, colors.white, "بارنامه های ثبت شده", 0, None, 
                      (None, 35, "bold"), 0.5, 0.1, "center")
        
        self.scrollable = scrollbar.ScrollableFrame(right_section_box, float(self.master.winfo_screenwidth() / 4.15), colors.dark_green_6)
        self.scrollable.pack(side="top", fill="x", padx=0, pady=(110, 0))
        
        for i, deal in enumerate(current_deals + old_deals):
            DealBox(parent, self.scrollable.inner_frame, middle_section_box, left_section_box, deal, float(self.master.winfo_screenwidth() / 4.4), 0.2 + i * 0.17)
        
        # لوگو
        ck.make_image(main_box, asset_paths.Logo_path, 88, 90, 0.02, 0.04, "nw")
        
        ck.make_button(main_box, "بازگشت", 150, 50, (None, 22, "bold"), 20, 
                       colors.white, colors.white, colors.dark_green_6, colors.green_3, 
                       self.go_to_view_bol, 0.075, 0.94, "center")
        
        
        
        
        