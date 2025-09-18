import bol_fake_data
import tkinter.ttk as ttk
import insystem_data
from PIL import Image, ImageDraw, ImageFont
from asset_paths import empty_bol
from bol_fake_data import current_bols
from tkinter.filedialog import asksaveasfilename
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import colors
import customtkinter as ctk
import common_controller as cc

def persian_to_eng_date(date_str):
    english_digits = '0123456789'
    persian_digits = '۰۱۲۳۴۵۶۷۸۹'
    trans_table = str.maketrans(''.join(persian_digits), ''.join(english_digits))
    if(date_str == ''): return
    return date_str.translate(trans_table)
    
def eng_to_persian_date(date_str):
    english_digits = '0123456789'
    persian_digits = '۰۱۲۳۴۵۶۷۸۹'
    trans_table = str.maketrans(''.join(english_digits), ''.join(persian_digits))
    if(date_str == '' or date_str == None): return 
    return date_str.translate(trans_table)

def cash_format(input):
    if(input != None and input != ''):
        rem = len(input) % 3
        text = ""
        if(rem): text = str((input[0:rem])) + ","
        for i in range(int(len(input[rem:]) / 3)): text += (input[rem:][3 * i : 3 * i + 3] + ",")
        return text[0 : len(text) - 1]
    return ''

def car_id_format_is_right(s):
    if len(s) != 14: return False
    persian_digits = '۰۱۲۳۴۵۶۷۸۹'
    digit_indexes = [0, 1, 3, 4, 5, 12, 13]
    for i in digit_indexes:
        if s[i] not in persian_digits: return False
    if not ('\u0600' <= s[2] <= '\u06FF' and s[2].isalpha()): return False
    if s[6] != '-': return False
    if s[7:12] != "ایران": return False
    return True
    
def car_id_true_format(car_id_str):
    if(car_id_str == ''): return
    return "ایران" + car_id_str[12:14] + "-" + car_id_str[3:6] + car_id_str[2] + car_id_str[0:2] 

def is_phone_format(phone):
    if(len(phone) != 11 or persian_to_eng_date(phone[0:2]) != "09"): return True
    return False
    
def does_item_exist(arr, input, target):
    for item in arr:
        if(persian_to_eng_date(item[input]) == persian_to_eng_date(target)): return item
    return None

def is_date_above(date, target):
    date_year = persian_to_eng_date(date[0:4])
    target_year = persian_to_eng_date(target[0:4])
    
    date_month = persian_to_eng_date(date[5:7])
    target_month = persian_to_eng_date(target[5:7])
    
    date_day = persian_to_eng_date(date[8:10])
    target_day = persian_to_eng_date(target[8:10])
    
    if(date_year > target_year): return True
    elif(date_year < target_year): return False
    else:
        if(date_month > target_month): return True
        elif(date_month < target_month): return False
        else: return date_day >= target_day

def get_group_items(arr, input, target):
    tmp = []
    for item in arr:
        if(eng_to_persian_date(arr[input]) == eng_to_persian_date(target)): tmp.append(item)
    return tmp
    
    
def get_package_old_deal(deal, package_num):
    bols = []
    for item in bol_fake_data.current_bols:
        if(eng_to_persian_date(item['deal'].split("/")[-1]) == eng_to_persian_date(deal["id"])
           and eng_to_persian_date(item['load_group_id']) == eng_to_persian_date(package_num)):
            bols.append(item)
    return bols

def get_bol_by_load_id(load_id):
    for bol in (bol_fake_data.current_bols + bol_fake_data.old_bols):
        if(eng_to_persian_date(bol["load_id"]) == eng_to_persian_date(load_id)): return bol
    return None

def match_input(input_str, target_str):
    input_str = persian_to_eng_date(input_str)
    target_str = persian_to_eng_date(target_str)
    for i in range(len(target_str)):
        if(input_str == target_str[0:i]): return True
    return False

def match_input2(input_str, target_str):
    if(input_str == ''): return True
    input_str = persian_to_eng_date(input_str)
    target_str = persian_to_eng_date(target_str)
    if(target_str == input_str): return False
    for i in range(len(target_str)):
        if(input_str == target_str[0:i]): return True
    return False

def get_driver_by_fullname(name):
    for driver in bol_fake_data.driver_sample_data:
        if(driver['name'] + " " + driver['lastname'] == name): return driver
    return None

def get_deal_bols(deal):
    bols = []
    for group in deal['packages']:
        bols += (group['bols'])
    return bols

def get_grouped_bols():
    bols = []
    for bol in bol_fake_data.old_bols:
        for deal in (insystem_data.current_deals + insystem_data.old_deals):
            if(eng_to_persian_date(bol['load_id']) in get_deal_bols(deal)): bols.append(bol)
    return bols

def get_deal_by_name_and_id(name, id):
    for item in insystem_data.current_deals:
        if(item['name'] == name and eng_to_persian_date(item['id']) == eng_to_persian_date(id)): return item
    return None

def get_load_type_id(load_name):
    for load in insystem_data.load_main_type:
        for i in range(len(load['sub_array'])):
            if(load['sub_array'][i] == load_name): return eng_to_persian_date(load['codes'][i])
    return None

def get_driver_licence(driver_name, driver_id):
    for driver in (bol_fake_data.driver_sample_data + bol_fake_data.old_drivers):
        if(driver['name'] + " " + driver['lastname'] == driver_name and 
           eng_to_persian_date(driver['id']) == eng_to_persian_date(driver_id)):
            return driver
    return None

def make_photo_with_data(bol):
    img = Image.open(empty_bol)

    draw = ImageDraw.Draw(img)
    small_font = ImageFont.truetype("Vazirmatn-SemiBold.ttf", size=18)
    font = ImageFont.truetype("Vazirmatn-SemiBold.ttf", size=30)
    middle_font = ImageFont.truetype("Vazirmatn-SemiBold.ttf", size=50)
    big_font = ImageFont.truetype("Vazirmatn-SemiBold.ttf", size=70)


    data = [
        (eng_to_persian_date(bol['cash_id']), (2130, 505)),
        (eng_to_persian_date(bol['load_id']), (2200, 410)),
        (eng_to_persian_date(bol['send_date']), (2200, 300)),
        (eng_to_persian_date(bol['send2_date']), (2200, 353)),
        (eng_to_persian_date(cash_format(bol['input_money'])), (1850, 300)),
        (eng_to_persian_date(cash_format(bol['payed_money'])), (1480, 300)),
        (eng_to_persian_date(bol['smart_cart_driver']), (1070, 300)),
        (eng_to_persian_date(bol['smart_cart_car']), (640, 300)),
        (eng_to_persian_date(bol['sent_type']), (300, 293)),
        (eng_to_persian_date(car_id_true_format(bol['car_id'])), (645, 350)),
        (eng_to_persian_date(bol['agent']), (700, 405)),
        (eng_to_persian_date(bol['driver_id']), (680, 460)),
        (eng_to_persian_date(bol['start_location']), (1565, 340)),
        (eng_to_persian_date(bol['des_location']), (1565, 400)),
        (eng_to_persian_date(cash_format(bol['true_value'])), (1670, 640)),
        (eng_to_persian_date(cash_format(bol['load_weight'])), (930, 640)),
    ]

    for text, pos in data:
        if(text != None and text != ''): draw.text(pos, text, (0,0,0), font=font, anchor='rt')

    draw.text((2200, 150), eng_to_persian_date(bol['deal']), (0, 0, 0), font=big_font, anchor='rt')
    draw.text((2200, 230), 'تاریخ ثبت بارنامه: ' + eng_to_persian_date(bol['today_date']), (0, 0, 0), font=middle_font, anchor='rt')
    draw.text((800, 200), eng_to_persian_date(bol['load_name']) + '/' + get_load_type_id(bol['load_name']), (0, 0, 0), font=middle_font, anchor='rt')
    draw.text((2275, 775), eng_to_persian_date(bol['driver_name']), (0, 0, 0), font=small_font, anchor='rt')
    draw.text((1927, 775), eng_to_persian_date((get_driver_licence(bol['driver_name'], bol['driver_id']))['licence_id']), (0, 0, 0), font=small_font, anchor='rt')
    
    file_path = asksaveasfilename(
        defaultextension=".png",
        filetypes=[
            ("PNG Image (*.png)", "*.png")
        ]
    )
    if file_path: img.save(file_path)

def make_string_arr(array):
    target = ''
    for i, item in enumerate(array):
        if(i == len(array) - 1): target += (item)
        else: target += (item + " - ")
    return target

def does_exist(input, arr):
    return input in arr

def get_car_by_carid(car_id):
    for car in bol_fake_data.car_sample_data:
        if(car['car_id'] == car_id): return car
    return None

def print_group(arr):
    wb = Workbook()
    ws = wb.active
    if(arr == bol_fake_data.driver_sample_data or arr == bol_fake_data.old_drivers):
        columns = ['ردیف', 'نام', 'نام خانوادگی', 'کد ملی', 'شماره تلفن', 'پلاک ماشین', 'شماره هوشمند', 'شماره گواهینامه', 'تاریخ شروع اعتبار هوشمند', 'تاریخ پایان اعتبار هوشمند']
    elif(arr == bol_fake_data.car_sample_data or arr == bol_fake_data.old_cars):
        columns = ['ردیف', 'پلاک ماشین', 'شماره هوشمند نفتکش', 'نام مالک', 'نام راننده', 'تاریخ شروع اعتبار هوشمند', 'تاریخ پایان اعتبار هوشمند', 'بنزین', 'نفت سفید', 'نفت گاز', 'نفت کوره']
    ws.append(columns)
    fill = PatternFill(start_color=colors.dark_green_6[1:], end_color=colors.dark_green_6[1:], fill_type="solid")
    font = Font(color=colors.white[1:], size=15,  bold=True)
    alignment = Alignment(horizontal="center", vertical="center")
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for i, item in enumerate(arr):
        values = list(item.values())
        tpl = (i + 1, ) + tuple(values)
        new_tpl = tuple(cc.eng_to_persian_date(str(x)) for x in tpl)
        ws.append(new_tpl)

    for col in range(1, 26):
        col_letter = ws.cell(row=1, column=col).column_letter
        ws.column_dimensions[col_letter].width = 15
    for col_num in range(1, len(columns)+1):
        ws.cell(row=1, column=col_num).fill = fill
        ws.cell(row=1, column=col_num).font = font
    for row_id in range(1 + len(arr)):
        ws.cell(row=row_id+1, column=1).fill = fill
        ws.cell(row=row_id+1, column=1).font = font
    for col_id in range(len(columns)):
        for row_id in range(1 + len(arr)):
            ws.cell(row=row_id+1, column=col_id+1).border = border
            ws.cell(row=row_id+1, column=col_id+1).alignment = alignment
    
    file_path = asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")]
    )
    if file_path:
        wb.save(file_path)
        

def activate_all_widgets(parent):
    for child in parent.winfo_children():
        if isinstance(child, ctk.CTkButton):
            child.configure(state='normal')
        if isinstance(child, ctk.CTkOptionMenu):
            child.configure(state="active")
        if isinstance(child, ctk.CTkEntry):
            child.configure(state='normal')
        if isinstance(child, ctk.CTkSlider):
            child.configure(state="normal")
        if isinstance(child, ctk.CTkSwitch):
            child.configure(state="normal")
        activate_all_widgets(child)
                
def disable_all_widgets(parent):
    for child in parent.winfo_children():
        if isinstance(child, ctk.CTkButton):
            child.configure(state='disabled')
        if isinstance(child, ctk.CTkOptionMenu):
            child.configure(state="disabled")
        if isinstance(child, ctk.CTkEntry):
            child.configure(state="disabled")
        if isinstance(child, ctk.CTkSlider):
            child.configure(state="disabled")
        if isinstance(child, ctk.CTkSwitch):
            child.configure(state="disabled")
        disable_all_widgets(child)

def find_manager():
    managers = []
    for user in insystem_data.users:
        if(user['role'] == 'مدیر'): managers.append(user)
    return managers