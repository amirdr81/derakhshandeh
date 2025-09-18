import customtkinter as ctk
import tkinter.ttk as ttk
import tkinter as tk
from bol_fake_data import current_bols, old_bols
import asset_paths
from date import Date
from insystem_data import load_main_type, old_deals, current_deals
from insystem_data import sent_type as types, agents
from bol_inf import BolInf
from tkinter.filedialog import asksaveasfilename
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from fpdf import FPDF
import register_controller as rc
import jdatetime
import common_controller as cc
import common_ctk as ck
import colors

class ReportPage(ctk.CTkFrame):    
    def update_bol_box(self, event):
            d_type = self.deal.get()
            b_type = self.load_type.get()
            a_type = self.agent.get()
            
            if(b_type == "بارنامه‌های فعال"): self.shown_data = current_bols
            elif(b_type == "بارنامه‌های آرشیو"): self.shown_data = old_bols
            elif(b_type == "همه‌ بارنامه‌ها"): self.shown_data = current_bols + old_bols
            else: self.shown_data = current_bols
            
            if(d_type != 'عدم وجود قرارداد' and d_type != 'همه قراردادها'): 
                self.shown_data = [bol for bol in self.shown_data if cc.eng_to_persian_date(bol["deal"]) == d_type]
            if(a_type != "همه نماینده‌ها"): 
                self.shown_data = [bol for bol in self.shown_data if cc.eng_to_persian_date(bol["agent"]) == a_type]
            
            
            self.update_table_with__new_data()
    
    def make_group_type2(self):
        label = ck.make_label(self, 150, 25, 
                          colors.white, colors.white, colors.red_color,
                          "", 0, None, (None, 13, "bold"), 0.9, 0.21, "e")
        selected_data = []
        for i, item_id in enumerate(self.tree.selection()):
            bol = cc.get_bol_by_load_id(self.tree.item(item_id, 'values')[1])
            selected_data.append(bol)
        if(len(selected_data) == 0):
            ck.update_label_error(label, 2000, 'بارنامه ای به جهت بستن، انتخاب نشده است!')
            return
        for item in selected_data:
            if(self.deal.get() == 'همه قراردادها'):
                ck.update_label_error(label, 2000, 'بارنامه ها باید برای یک قرارداد باشند!')
                return
        for item in selected_data:
            if(self.load_type.get() != 'بارنامه‌های فعال'):
                ck.update_label_error(label, 2000, 'بستن صورت، باید از میان بارنامه های فعال باشد!')
                return
            
        def do_action():
            deal_name, deal_id = self.deal.get().split('/')
            deal = cc.get_deal_by_name_and_id(deal_name, deal_id)
            num_of_groups = len(deal['packages'])
            if(num_of_groups == 0): last_group = None
            else: last_group = deal['packages'][num_of_groups - 1]
            bols = []
            
            for bol in selected_data:
                if(num_of_groups == 0): bol['load_group_id'] = '۱'
                else: bol['load_group_id'] = cc.eng_to_persian_date(str(1 + int(last_group['id'])))
                bols.append(cc.eng_to_persian_date(bol['load_id']))
                old_bols.append(bol)
                current_bols.remove(bol)
            
            if(num_of_groups == 0):
                (deal['packages']).append({
                'id': '۱',
                'bols': bols,
                'submit_date': cc.eng_to_persian_date(jdatetime.date.today().strftime('%Y/%m/%d')),
                'payed': False
            })
            else:    
                (deal['packages']).append({
                    'id': cc.eng_to_persian_date(str(1 + int(last_group['id']))),
                    'bols': bols,
                    'submit_date': cc.eng_to_persian_date(jdatetime.date.today().strftime('%Y/%m/%d')),
                    'payed': False
                })
            
            for row_id in self.tree.get_children():
                self.tree.delete(row_id)
            cancel()
            self.go_to_bol_groups()
        def cancel():
            cc.activate_all_widgets(self)
            main_frame.destroy()
        cc.disable_all_widgets(self)
        main_frame = rc.show_confirmation_box(self, do_action, cancel, 'آیا از بستن این صورت، اطمینان دارید؟', None)
        
    def make_group(self):
        label = ck.make_label(self, 150, 25, 
                          colors.white, colors.white, colors.red_color,
                          "", 0, None, (None, 13, "bold"), 0.9, 0.21, "e")
        if(self.deal.get() == 'عدم وجود قرارداد' or self.deal.get() == 'همه قراردادها'): 
            ck.update_label_error(label, 2000, 'برای بستن صورت، حتما باید یک قرارداد را انتخاب کرده باشید!')
        elif(self.load_type.get() != 'بارنامه‌های فعال'):
            ck.update_label_error(label, 2000, 'بستن صورت، باید از میان بارنامه های فعال باشد!')
        elif(len(self.tree.get_children()) == 0):
            ck.update_label_error(label, 2000, 'بارنامه‌ای به جهت بستن صورت، وجود ندارد!')
        else:
            def do_action():
                deal_name, deal_id = self.deal.get().split('/')
                deal = cc.get_deal_by_name_and_id(deal_name, deal_id)
                num_of_groups = len(deal['packages'])
                if(num_of_groups == 0): last_group = None
                else: last_group = deal['packages'][num_of_groups - 1]
                bols = []
                
                for row_id in self.tree.get_children():
                    bol = cc.get_bol_by_load_id(cc.eng_to_persian_date(str(self.tree.item(row_id)['values'][1])))
                    if(num_of_groups == 0): bol['load_group_id'] = '۱'
                    else: bol['load_group_id'] = cc.eng_to_persian_date(str(1 + int(last_group['id'])))
                    bols.append(cc.eng_to_persian_date(bol['load_id']))
                    old_bols.append(bol)
                    current_bols.remove(bol)
                
                if(num_of_groups == 0):
                    (deal['packages']).append({
                    'id': '۱',
                    'bols': bols,
                    'submit_date': cc.eng_to_persian_date(jdatetime.date.today().strftime('%Y/%m/%d')),
                    'payed': False
                })
                else:    
                    (deal['packages']).append({
                        'id': cc.eng_to_persian_date(str(1 + int(last_group['id']))),
                        'bols': bols,
                        'submit_date': cc.eng_to_persian_date(jdatetime.date.today().strftime('%Y/%m/%d')),
                        'payed': False
                    })
                
                for row_id in self.tree.get_children():
                    self.tree.delete(row_id)
                cancel()
                
                self.go_to_bol_groups()
                
            def cancel():
                cc.activate_all_widgets(self)
                frame.destroy()
            cc.disable_all_widgets(self)
            frame = rc.show_confirmation_box(self, do_action, cancel, 
                                     'آیا از بستن صورت، اطمینان دارید؟', 
                                     None)
            
    def get_excel_output(self):
        wb = Workbook()
        ws = wb.active
        columns = ['ردیف', 'تاریخ ثبت', 'قرارداد', 'شماره حواله', 'شماره بارنامه', 'شماره صورت بارنامه', 'نماینده', 'شماره صورت نماینده', 'تاریخ صدور', 'تاریخ ارسال', 'نام راننده', 'کد ملی راننده', 'شماره هوشمند راننده', 'نام مالک', 'شماره نفتکش', 'شماره هوشمند نفتکش', 'مبدأ', 'مقصد', 'وزن واقعی', 'وزن حمل شده', 'عنوان کالا', 'نام فرآورده', 'نوع ارسال', 'مقدار دریافتی', 'مقدار پرداختی']
        ws.append(columns)
        counter = 1
        fill = PatternFill(start_color=colors.dark_green_6[1:], end_color=colors.dark_green_6[1:], fill_type="solid")
        font = Font(color=colors.white[1:], size=15,  bold=True)
        alignment = Alignment(horizontal="center", vertical="center")
        thin = Side(border_style="thin", color="000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        
        true_value_sum = 0
        load_weight_sum = 0
        input_money_sum = 0
        payed_money_sum = 0
        
        for item in self.shown_data:
            for row_data in self.tree.get_children():
                if (self.tree.item(row_data, 'values')[1] == item["load_id"]):
                    new_list = list(item.values())
                    true_value_sum += int(cc.persian_to_eng_date(new_list[17]))
                    load_weight_sum += int(cc.persian_to_eng_date(new_list[18]))
                    input_money_sum += int(cc.persian_to_eng_date(new_list[22]))
                    payed_money_sum += int(cc.persian_to_eng_date(new_list[23]))
                    new_list[17] = cc.cash_format(new_list[17])
                    new_list[18] = cc.cash_format(new_list[18])
                    new_list[22] = cc.cash_format(new_list[22])
                    new_list[23] = cc.cash_format(new_list[23])
                    tpl = (counter, ) + tuple(new_list)
                    new_tpl = tuple(cc.eng_to_persian_date(str(x)) for x in tpl)
                    ws.append(new_tpl)
                    counter += 1
        ws.cell(row=counter + 1, column=19, value=cc.eng_to_persian_date(cc.cash_format(str(true_value_sum))))
        ws.cell(row=counter + 1, column=20, value=cc.eng_to_persian_date(cc.cash_format(str(load_weight_sum))))
        ws.cell(row=counter + 1, column=24, value=cc.eng_to_persian_date(cc.cash_format(str(input_money_sum))))
        ws.cell(row=counter + 1, column=25, value=cc.eng_to_persian_date(cc.cash_format(str(payed_money_sum))))
        
        for col in range(1, 26):
            col_letter = ws.cell(row=1, column=col).column_letter
            if col == 1: ws.column_dimensions[col_letter].width = 8
            if col in [19, 20, 21, 22, 23, 25, 26]: ws.column_dimensions[col_letter].width = 15
            if col in [2, 3, 4, 5, 6, 8, 9, 10, 11, 12, 13, 14, 15, 16, 24]: ws.column_dimensions[col_letter].width = 20
            if col in [7, 17, 18]: ws.column_dimensions[col_letter].width = 30
        for col_num in range(1, len(columns)+1):
            ws.cell(row=1, column=col_num).fill = fill
            ws.cell(row=1, column=col_num).font = font
        for row_id in range(1 + len(self.tree.get_children())):
            ws.cell(row=row_id+1, column=1).fill = fill
            ws.cell(row=row_id+1, column=1).font = font
        for col_id in range(len(columns)):
            for row_id in range(1 + len(self.tree.get_children())):
                ws.cell(row=row_id+1, column=col_id+1).border = border
                ws.cell(row=row_id+1, column=col_id+1).alignment = alignment
        
        ws.cell(row=1 + counter, column=19).fill = fill
        ws.cell(row=1 + counter, column=19).font = font
        ws.cell(row=1 + counter, column=20).fill = fill
        ws.cell(row=1 + counter, column=20).font = font
        ws.cell(row=1 + counter, column=24).fill = fill
        ws.cell(row=1 + counter, column=24).font = font
        ws.cell(row=1 + counter, column=25).fill = fill
        ws.cell(row=1 + counter, column=25).font = font
        
        ws.cell(row=1 + counter, column=19).border = border
        ws.cell(row=1 + counter, column=19).alignment = alignment
        ws.cell(row=1 + counter, column=20).border = border
        ws.cell(row=1 + counter, column=20).alignment = alignment
        ws.cell(row=1 + counter, column=24).border = border
        ws.cell(row=1 + counter, column=24).alignment = alignment
        ws.cell(row=1 + counter, column=25).border = border
        ws.cell(row=1 + counter, column=25).alignment = alignment
        
        file_path = asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if file_path:
            wb.save(file_path)
        
    def find_initial_result(self, text):
        sum = 0
        for item in self.shown_data: sum += int(self.persian_to_eng_num(item[text]))
        return str(sum)
    
    def update_result_box(self, result1, result2, result3, result4, result5, result6):
        self.sum_weights_litr.configure(text=f"جمع وزن حمل شده(لیتر): {self.cash_format(self.eng_to_persian_num(str(result1)))}")
        self.sum_weights_kilo.configure(text=f"جمع وزن حمل شده(کیلو): {self.cash_format(self.eng_to_persian_num(str(result2)))}")
        self.sum_weights_litr2.configure(text=f"جمع وزن واقعی (لیتر): {self.cash_format(self.eng_to_persian_num(str(result3)))}")
        self.sum_weights_kilo2.configure(text=f"جمع وزن واقعی(کیلو): {self.cash_format(self.eng_to_persian_num(str(result4)))}")
        self.sum_cash_input.configure(text=f"جمع کرایه دریافتی: {self.cash_format(self.eng_to_persian_num(str(result5)))}")
        self.sum_cash_output.configure(text=f"جمع کرایه پرداختی: {self.cash_format(self.eng_to_persian_num(str(result6)))}")
        
    def get_sum_of_dataset(self, dataset, text):
        sum = 0
        for item in dataset: sum += int(self.persian_to_eng_num(item[text]))
        return sum
        
    def find_endpoints_value(self, text_input):
        if(len(self.shown_data) == 0): return 0, 1
        min_value = int(self.persian_to_eng_num(self.shown_data[0][text_input]))
        max_value = int(self.persian_to_eng_num(self.shown_data[0][text_input]))
        for item in self.shown_data:
            if(int(self.persian_to_eng_num(item[text_input])) < min_value): min_value = int(self.persian_to_eng_num(item[text_input]))
            if(int(self.persian_to_eng_num(item[text_input])) > max_value): max_value = int(self.persian_to_eng_num(item[text_input]))
        if(min_value == max_value): return min_value, max_value + 1
        return min_value, max_value
        
    def is_cash_above(self, cash, target):
        return (cash >= target)
        
    def check_range_cash(self, start_cash, end_cash, is_active, target_cash):
        start_cash = int(start_cash)
        end_cash = int(end_cash)
        target_cash = int(self.persian_to_eng_num(target_cash))
        if(not is_active): return self.is_cash_above(target_cash, start_cash)
        else: return (self.is_cash_above(target_cash, start_cash) and self.is_cash_above(end_cash, target_cash))
    
    def is_date(self, date_str):
        return (len(date_str) == 10 and self.is_num(date_str[0,1,2,3,5,6,8,9]))
        
    def is_date_above(self, date, target):
        date_year = self.persian_to_eng_num(date[0:4])
        target_year = self.persian_to_eng_num(target[0:4])
        
        date_month = self.persian_to_eng_num(date[5:7])
        target_month = self.persian_to_eng_num(target[5:7])
        
        date_day = self.persian_to_eng_num(date[8:10])
        target_day = self.persian_to_eng_num(target[8:10])
        
        if(date_year > target_year): return True
        elif(date_year < target_year): return False
        else:
            if(date_month > target_month): return True
            elif(date_month < target_month): return False
            else: return date_day >= target_day
        
    def check_range_date(self, start_date, end_date, is_active, target_date):
        if(not is_active): return start_date == target_date
        else: return (self.is_date_above(target_date, start_date) and self.is_date_above(end_date, target_date))
           
    def is_num(self, str):
        if(len(str) == 0): return False
        for i in range(len(str)):
            if(str[i] not in ['0', '1', '2', '3', '4', '5', '6', '7', '8', '9', '۰', '۱', '۲', '۳', '۴', '۵', '۶', '۷', '۸', '۹']): return False
        return True
    def check_range_load_id(self, start_point, end_point, is_active, target_point):
        if(not is_active): return self.match_input((start_point), (target_point))
        if(self.is_num(start_point) and self.is_num(end_point) and self.is_num(target_point)): 
            start_point = int(self.persian_to_eng_num(start_point))
            end_point = int(self.persian_to_eng_num(end_point))
            target_point = int(self.persian_to_eng_num(target_point))
            return ((start_point <= target_point) and (target_point <= end_point))
        return False
    
    def match_input(self, input_str, target_str):
        input_str = self.persian_to_eng_num(input_str)
        target_str = self.persian_to_eng_num(target_str)
        def is_in_first_part(target):
            for i in range(1 + len(target)):
                if(input_str == target[0:i]): return True
            return False
        words = target_str.split(" ")
        for word in words:
            if(is_in_first_part(word)): return True
        return False
    
    def cash_format(self, input):
        rem = len(input) % 3
        text = ""
        if(rem): text = str((input[0:rem])) + ","
        for i in range(int(len(input[rem:]) / 3)): text += (input[rem:][3 * i : 3 * i + 3] + ",")
        return text[0 : len(text) - 1]
        
    def toggle_mode_real_gone(self, label, is_input):
        if is_input.get(): 
            label.configure(text="محموله")
            self.load_slider1.configure(from_=self.find_endpoints_value("load_weight")[0], to=self.find_endpoints_value("load_weight")[1])
            self.load_slider2.configure(from_=self.find_endpoints_value("load_weight")[0], to=self.find_endpoints_value("load_weight")[1])
            self.load_slider1.set(self.find_endpoints_value("load_weight")[0])
            self.load_slider2.set(self.find_endpoints_value("load_weight")[1])
            self.load_slider1_label.configure(text=cc.eng_to_persian_date(self.cash_format(f"{int(self.load_slider1.get())}")))
            self.load_slider2_label.configure(text=cc.eng_to_persian_date(self.cash_format(f"{int(self.load_slider2.get())}")))
        else: 
            label.configure(text="واقعی")
            self.load_slider1.configure(from_=self.find_endpoints_value("true_value")[0], to=self.find_endpoints_value("true_value")[1])
            self.load_slider2.configure(from_=self.find_endpoints_value("true_value")[0], to=self.find_endpoints_value("true_value")[1])
            self.load_slider1.set(self.find_endpoints_value("true_value")[0])
            self.load_slider2.set(self.find_endpoints_value("true_value")[1])
            self.load_slider1_label.configure(text=cc.eng_to_persian_date(self.cash_format(f"{int(self.load_slider1.get())}")))
            self.load_slider2_label.configure(text=cc.eng_to_persian_date(self.cash_format(f"{int(self.load_slider2.get())}")))
            
        self.update_table_with__new_data()
    def toggle_mode_input_output(self, label, is_input):
        if is_input.get(): 
            label.configure(text="پرداختی")
            self.cash_slider1.configure(from_=self.find_endpoints_value("payed_money")[0], to=self.find_endpoints_value("payed_money")[1])
            self.cash_slider2.configure(from_=self.find_endpoints_value("payed_money")[0], to=self.find_endpoints_value("payed_money")[1])
            self.cash_slider1.set(self.find_endpoints_value("payed_money")[0])
            self.cash_slider2.set(self.find_endpoints_value("payed_money")[1])
            self.cash_slider1_label.configure(text=cc.eng_to_persian_date(self.cash_format(f"{int(self.cash_slider1.get())}")))
            self.cash_slider2_label.configure(text=cc.eng_to_persian_date(self.cash_format(f"{int(self.cash_slider2.get())}")))
        else: 
            self.cash_slider1.configure(from_=self.find_endpoints_value("input_money")[0], to=self.find_endpoints_value("input_money")[1])
            self.cash_slider2.configure(from_=self.find_endpoints_value("input_money")[0], to=self.find_endpoints_value("input_money")[1])
            self.cash_slider1.set(self.find_endpoints_value("input_money")[0])
            self.cash_slider2.set(self.find_endpoints_value("input_money")[1])
            self.cash_slider1_label.configure(text=cc.eng_to_persian_date(self.cash_format(f"{int(self.cash_slider1.get())}")))
            self.cash_slider2_label.configure(text=cc.eng_to_persian_date(self.cash_format(f"{int(self.cash_slider2.get())}")))
            label.configure(text="دریافتی")
        self.update_table_with__new_data()
    def toggle_mode_load_slider(self, entry2, entry2_label, entry_value, entry2_value, is_range, yrel,text):
        if is_range.get():
            entry_value.configure(text=f"از {text}")
            entry2.place(relx=0.55, rely=yrel, anchor="center")
            entry2_label.place(relx=0.35, rely=yrel, anchor="center")
            entry2_value.place(relx=0.745, rely=yrel, anchor="center")
        else:
            entry_value.configure(text=f"{text}")
            entry2.place_forget()
            entry2_label.place_forget()
            entry2_value.place_forget()
        self.update_table_with__new_data()
        
    def toggle_mode_cash_slider(self, entry2, entry2_label, entry_value, entry2_value, is_range, yrel,text):
        if is_range.get():
            entry_value.configure(text=f"از {text}")
            entry2.place(relx=0.6, rely=yrel, anchor="center")
            entry2_label.place(relx=0.369, rely=yrel, anchor="center")
            entry2_value.place(relx=0.79, rely=yrel, anchor="center")
        else:
            entry_value.configure(text=f"{text}")
            entry2.place_forget()
            entry2_label.place_forget()
            entry2_value.place_forget()
        self.update_table_with__new_data()
            
    def toggle_mode_load_id(self, entry1, entry2, is_range, xrel, yrel, text):
        if is_range.get():
            entry1.configure(placeholder_text=f"از {text}", state="normal")
            entry2.configure(placeholder_text=f"تا {text}", state="normal")
            entry2.place(relx=xrel, rely=yrel, anchor="center")
        else:
            entry1.configure(placeholder_text=text, state="normal")
            entry2.delete(0, "end")
            entry2.place_forget()
        self.update_table_with__new_data()
            
    def toggle_mode(self, entry1, entry2, is_range, xrel, yrel, text):
        if is_range.get():
            entry1.configure(text=f"از {text}", text_color=colors.gray_3, state="normal")
            entry2.configure(text=f"تا {text}", fg_color=colors.dark_gray_color, text_color=colors.light_gray_7, state="disabled")
            entry2.place(relx=xrel, rely=yrel, anchor="center")
        else:
            entry1.configure(text=text, text_color=colors.gray_3, state="normal")
            entry2.place_forget()
        self.update_table_with__new_data()

    def get(self, entry1, entry2, is_range):
        val1 = entry1.get()
        val2 = entry2.get() if is_range.get() else None
        return (val1, val2)
    
    def persian_to_eng_num(self, date_str):
        english_digits = '0123456789'
        persian_digits = '۰۱۲۳۴۵۶۷۸۹'
        trans_table = str.maketrans(''.join(persian_digits), ''.join(english_digits))
        return date_str.translate(trans_table)
    
    def eng_to_persian_num(self, date_str):
        english_digits = '0123456789'
        persian_digits = '۰۱۲۳۴۵۶۷۸۹'
        trans_table = str.maketrans(''.join(english_digits), ''.join(persian_digits))
        return date_str.translate(trans_table)
    
    #آپدیت شدن جدول با داده های جدید
    def update_table_with__new_data(self, event=None):
        result1 = 0
        result2 = 0
        result3 = 0
        result4 = 0
        result5 = 0
        result6 = 0
        for child in self.tree.get_children(): self.tree.delete(child)
        flag = 0
        for item in self.shown_data:
            if((self.main_type_list.get() == "عنوان کالا" or item["main_type"] == self.main_type_list.get()) and
                (self.load_name.get() == "فرآورده" or item["load_name"] == self.load_name.get()) and
                (self.sent_type.get() == "نوع ارسال" or item["sent_type"] == self.sent_type.get()) and
                (not self.driver_name_entry.get() or self.match_input(self.driver_name_entry.get(), item["driver_name"])) and
                (not self.driver_id_entry.get() or self.match_input(self.persian_to_eng_num(self.driver_id_entry.get()), self.persian_to_eng_num(item["driver_id"]))) and
                (not self.owner_name_entry.get() or self.match_input(self.owner_name_entry.get(), item["car_owner_name"])) and
                (not self.driver_smart_id_entry.get() or self.match_input(self.driver_smart_id_entry.get(), item["smart_cart_driver"])) and
                (not self.licence_safety_end.get() or self.match_input(self.licence_safety_end.get(), item["car_id"])) and
                (not self.car_smart_id_entry.get() or self.match_input(self.car_smart_id_entry.get(), item["smart_cart_car"])) and
                (not self.bol_first_entry.get() or self.check_range_load_id(self.bol_first_entry.get(), self.bol_second_entry.get(), self.bol_switch_bin.get(), item["load_id"])) and
                ((self.register_date_btn.cget("text") == "تاریخ ثبت") or (self.check_range_date(self.register_date_btn.cget("text"), self.register_date_btn2.cget("text"), self.register_date_bin.get(), item["today_date"]))) and
                (self.sent_date_btn.cget("text") == "تاریخ صدور" or self.check_range_date(self.sent_date_btn.cget("text"), self.sent_date_btn2.cget("text"), self.date_switch_bin.get(), item["send_date"])) and
                (self.sent2_date_btn.cget("text") == "تاریخ ارسال" or self.check_range_date(self.sent2_date_btn.cget("text"), self.sent2_date_btn2.cget("text"), self.register_date_bin2.get(), item["send2_date"])) and
                (not self.start_route_entry.get() or self.match_input(self.start_route_entry.get(), item["start_location"])) and
                (not self.end_route_entry.get() or self.match_input(self.end_route_entry.get(), item["des_location"])) and
                ((self.input_output_bin.get() and self.check_range_cash(self.cash_slider1.get(), self.cash_slider2.get(), self.cash_switch_bin.get(), item["payed_money"])) or 
                 (not self.input_output_bin.get() and self.check_range_cash(self.cash_slider1.get(), self.cash_slider2.get(), self.cash_switch_bin.get(), item["input_money"]))) and
                ((self.real_gone_bin.get() and self.check_range_cash(self.load_slider1.get(), self.load_slider2.get(), self.load_switch_bin.get(), item["load_weight"])) or 
                 (not self.real_gone_bin.get() and self.check_range_cash(self.load_slider1.get(), self.load_slider2.get(), self.load_switch_bin.get(), item["true_value"])))
                ):
                
                self.add_row((
                            cc.eng_to_persian_date(str(flag + 1)),
                            cc.eng_to_persian_date(item["load_id"]),
                            cc.eng_to_persian_date(item["load_name"]),
                            cc.eng_to_persian_date(item["start_location"]),
                            cc.eng_to_persian_date(item["des_location"]),
                            cc.eng_to_persian_date(cc.car_id_true_format(item["car_id"])),
                            cc.eng_to_persian_date(item["load_weight"])
                ))
                if(item["load_weight"] != '' and item["load_weight"] != None):
                    result1 += int(self.persian_to_eng_num(item["load_weight"]))
                    result2 += int(self.persian_to_eng_num(item["load_weight"]))
                if(item["true_value"] != '' and item["true_value"] != None):
                    result3 += int(self.persian_to_eng_num(item["true_value"]))
                    result4 += int(self.persian_to_eng_num(item["true_value"]))
                if(item["input_money"] != '' and item["input_money"] != None):
                    result5 += int(self.persian_to_eng_num(item["input_money"]))
                if(item["payed_money"] != '' and item["payed_money"] != None):
                    result6 += int(self.persian_to_eng_num(item["payed_money"]))
                flag += 1
        self.update_result_box(result1, result2, result3, result4, result5, result6)
        
    def add_row(self, data_tuple):
        iid = self.tree.insert("", tk.END, values=data_tuple)
        self.tree.tag_configure(f"hover_{iid}", background=colors.light_green_5)
        self.tree.item(iid, tags=(f"hover_{iid}",))
        return iid   
    
    def car_id_true_format(self, car_id_str):
        return "ایران " + car_id_str[12:14] + " - " + car_id_str[3:6] + " " + car_id_str[2] + " " + car_id_str[0:2] 
        
    def __init__(self, parent, go_to_dashboard, go_to_veiw_bol, go_to_bol_groups, reset_page):
        super().__init__(parent)
        self.go_to_dashboard = go_to_dashboard
        self.pack(fill="both", expand=True)
        self.configure(fg_color=colors.light_green_1)
        self.go_to_veiw_bol = go_to_veiw_bol          
        self.go_to_bol_groups = go_to_bol_groups
        self.reset_page = reset_page
        self.shown_data = current_bols
        
        def update_live_information(entry):
            entry.bind("<KeyRelease>", lambda event: self.update_table_with__new_data(event=event))
            entry.bind("<FocusOut>", lambda event: self.update_table_with__new_data(event=event))
            
        def activate_all_widgets(parent):
            for child in parent.winfo_children():
                if isinstance(child, ctk.CTkButton):
                    child.configure(state='normal')
                activate_all_widgets(child)
                     
        def disable_all_widgets(parent):
            for child in parent.winfo_children():
                if isinstance(child, ctk.CTkButton):
                    child.configure(state='disabled')
                    
                disable_all_widgets(child)
                
        def open_bol():
            global bol_frame
            try: bol_frame.destroy()
            except: pass
            bol_frame = ck.make_frame(self, float(self.master.winfo_screenwidth() / 1.3), float(self.master.winfo_screenheight() / 2.5),
                                      colors.light_green_1, colors.light_green_1, colors.light_green_1, 0, 0, 0.5, 0.5, "center")

            cal = BolInf(bol_frame, reset_page, cc.get_bol_by_load_id(self.tree.item(self.tree.selection()[0], 'values')[1]))
            cal.place(relx=0.5, rely=0.5, anchor="center")
            ck.make_button(bol_frame, 'بستن', 120, 30, (None, 25), 60, colors.light_green_1,
                           colors.light_green_1, colors.dark_green_6, colors.green_3,
                           lambda: bol_frame.destroy(), 0.075, 0.91, "center")
            
        def open_calendar(box, entry, x_co, y_co):
            global calendar_frame
            try: calendar_frame.destroy()
            except: pass
            calendar_frame = tk.Frame(box, bd=2, relief=tk.RIDGE,
                                      bg=colors.light_green_1)
            calendar_frame.place(relx=x_co, rely=y_co, anchor="center")
            def on_date_selected(selected_date):
                if(self.register_date_btn.cget("text") and self.register_date_btn2.cget("state") == "disabled"): self.register_date_btn2.configure(state="normal", fg_color=colors.white, text_color=colors.gray_3)
                if(self.sent_date_btn.cget("text") and self.sent_date_btn2.cget("state") == "disabled"): self.sent_date_btn2.configure(state="normal", fg_color=colors.white, text_color=colors.gray_3)
                if(self.sent2_date_btn.cget("text") and self.sent2_date_btn2.cget("state") == "disabled"): self.sent2_date_btn2.configure(state="normal", fg_color=colors.white, text_color=colors.gray_3)
                entry.configure(text=str(selected_date).replace("-", "/"), text_color=colors.black)
                self.update_table_with__new_data()
                calendar_frame.destroy()

            cal = Date(calendar_frame, callback=on_date_selected)
            cal.pack()

            close_btn = ctk.CTkButton(calendar_frame, text="بستن", 
                                      width=100,
                                      bg_color=colors.light_green_1, 
                                fg_color=colors.dark_green_6,
                                text_color=colors.light_green_1,
                                hover_color=colors.green_3,
                                command=calendar_frame.destroy)
            close_btn.pack(pady=4)
        
        def make_box(box,width, height, bg_color, fg_color, corner_radius, x_co, y_co):
            made_box = ctk.CTkFrame(box, width=width, 
                                height=height, 
                                bg_color=bg_color,
                                fg_color=fg_color,
                                corner_radius=corner_radius)
            made_box.place(relx=x_co, rely=y_co, anchor="center")
            return made_box
        
        def make_entry_box(box, text, height, width, bg_color, fg_color, border_color, text_color, x_co, y_co, state):
            entry = ctk.CTkEntry(box, 
                                placeholder_text=text, 
                                height=height , justify="right", width=width, 
                                corner_radius=60, 
                                bg_color=bg_color, 
                                state=state,
                                border_color=border_color, 
                                fg_color=fg_color, 
                                text_color=text_color)
            if(x_co and y_co): entry.place(relx=x_co, rely=y_co, anchor="center")
            return entry
        
        main_box = make_box(self, float(self.master.winfo_screenwidth() / 1.2), float(self.master.winfo_screenheight() / 1.5), colors.light_green_1, colors.white, 0, 0.5, 0.5)
        top_left_box = make_box(main_box, float(self.master.winfo_screenwidth() / 2.5), float(self.master.winfo_screenheight() / 2.5), colors.white, colors.dark_green_6, 20, 0.27, 0.48)
        buttom_left_box = make_box(main_box, float(self.master.winfo_screenwidth() / 2.5), float(self.master.winfo_screenheight() / 17), colors.white, colors.light_green_4, 20, 0.27, 0.835)
        
        self.column_titles = ["ردیف", "شماره بارنامه", "نوع فرآورده", "مبدأ", "مقصد", "شماره نفتکش", "وزن محموله"]
        self.column_ids = [f"col{i}" for i in range(len(self.column_titles))]

        # استایل Treeview
        style = ttk.Style(self)
        style.theme_use("default")
        style.configure("Treeview", background=colors.light_green_5,
                        foreground=colors.black, fieldbackground=colors.light_green_4, rowheight=30)
        style.configure("Treeview.Heading", background=colors.dark_green_6,
                        foreground=colors.white, font=(None, 15, "bold"),
                        padding=(5, 5))
        # رنگ ردیف انتخاب شده
        style.map("Treeview",
                  background=[("selected", colors.green_2)],
                  foreground=[("selected", "white")])

        style.configure("Custom.Vertical.TScrollbar", 
            background=colors.dark_green_6,
            troughcolor=colors.light_green_1,
            bordercolor=colors.white,
            arrowcolor=colors.white,
            relief="flat")
        style.map("Custom.Vertical.TScrollbar",
            background=[("active", colors.green_2)],         # وقتی ماوس روی اسکرول‌بار، رنگ قرمز
            arrowcolor=[("active", colors.white)]
        )

        style.configure("Custom.Horizontal.TScrollbar", 
            background=colors.dark_green_6,
            troughcolor=colors.light_green_1,
            bordercolor=colors.white,
            arrowcolor=colors.white,
            relief="flat")
        style.map("Custom.Horizontal.TScrollbar",
            background=[("active", colors.green_2)],         # وقتی ماوس روی اسکرول‌بار، رنگ قرمز
            arrowcolor=[("active", colors.white)]
        )
        
        table_width = main_box.winfo_screenwidth() / 2.6
        table_height = main_box.winfo_screenheight() / 1.9
        scroll_bar_width = 16
        
        # Frame والد با ابعاد دقیقاً ۲۰۰x۱۰۰
        tree_frame = ctk.CTkFrame(main_box, width=table_width, height=table_height, fg_color=colors.white)
        tree_frame.place(relx=0.757, rely=0.58, anchor="center")
        tree_frame.pack_propagate(False)

        # Treeview با ابعاد کاملاً فیکس
        self.tree = ttk.Treeview(
            tree_frame,
            columns=self.column_ids,
            show="headings",
            height=4,   # صرفاً ارتفاع پیش‌فرض ردیف‌ها برای دید اولیه، تاثیری در ابعاد نهایی ندارد
            selectmode="extended"
        )
        self.tree.place(x=0, y=0, width=table_width, height=table_height)

        tree_scrollbar_y = ttk.Scrollbar(tree_frame, 
                                         orient="vertical", 
                                         command=self.tree.yview, 
                                         style="Custom.Vertical.TScrollbar")
        tree_scrollbar_y.place(x=table_width-scroll_bar_width, y=0, width=scroll_bar_width, height=table_height)
        
        scroll_x = ttk.Scrollbar(tree_frame, 
                                 orient="horizontal", 
                                 command=self.tree.xview,
                                 style="Custom.Horizontal.TScrollbar")
        scroll_x.place(x=0, y=table_height-scroll_bar_width, width=table_width-scroll_bar_width, height=scroll_bar_width)

        self.tree.configure(yscrollcommand=tree_scrollbar_y.set, xscrollcommand=scroll_x.set)

        # هر ستون غیرقابل استرچ، و با عرض کوچک (اگر زیاد بودن، فقط با اسکرول ظاهر می‌شوند)
        for i, name in enumerate(self.column_titles):
            if i == 0:
                self.tree.heading(0, text=name, anchor=tk.CENTER)
                self.tree.column(0, width=50, minwidth=50, anchor=tk.CENTER, stretch=tk.NO) 
            elif i == 4:
                self.tree.heading(4, text=name, anchor=tk.CENTER)
                self.tree.column(4, width=350, minwidth=50, anchor=tk.CENTER, stretch=tk.NO) 
            elif i == 6:
                self.tree.heading(6, text=name, anchor=tk.CENTER)
                self.tree.column(6, width=120, minwidth=50, anchor=tk.CENTER, stretch=tk.NO) 
            else:
                self.tree.heading(i, text=name, anchor=tk.CENTER)
                self.tree.column(i, width=140, minwidth=50, anchor=tk.CENTER, stretch=tk.NO)
        # داده‌ها
        for i, item in enumerate(self.shown_data):
            self.add_row((
                            cc.eng_to_persian_date(str(i + 1)),
                            cc.eng_to_persian_date(item["load_id"]),
                            cc.eng_to_persian_date(item["load_name"]),
                            cc.eng_to_persian_date(item["start_location"]),
                            cc.eng_to_persian_date(item["des_location"]),
                            cc.eng_to_persian_date(cc.car_id_true_format(item["car_id"])),
                            cc.eng_to_persian_date(item["load_weight"])
                ))
        
        # هوور ساده
        self._last_hovered = None
        def on_tree_motion(event):
            row = self.tree.identify_row(event.y)
            if row != self._last_hovered:
                if self._last_hovered is not None and self.tree.exists(self._last_hovered):
                    self.tree.tag_configure(f"hover_{self._last_hovered}", background=colors.light_green_5)
                if row and self.tree.exists(row):
                    self.tree.tag_configure(f"hover_{row}", background=colors.green_2)
                self._last_hovered = row
        def on_tree_leave(event):
            if self._last_hovered and self.tree.exists(self._last_hovered):
                self.tree.tag_configure(f"hover_{self._last_hovered}", background=colors.light_green_5)
            self._last_hovered = None

        self.tree.bind("<Motion>", on_tree_motion)
        self.tree.bind("<Leave>", on_tree_leave)

        # عنوان
        sign_in_label = ctk.CTkLabel(main_box, text="گزارش گیری", font=(None, 45, "bold"), text_color=colors.black, bg_color=colors.white)
        sign_in_label.place(relx=0.28, rely=0.1, anchor="center")
        
        #چاپ
        print_result = ctk.CTkButton(main_box, text="چاپ", 
                                   width=100, height=50, 
                                   font=(None, 22, "bold"), 
                                   corner_radius=20, 
                                   text_color=colors.white, 
                                   bg_color=colors.white, 
                                   fg_color=colors.dark_green_5, 
                                   hover_color=colors.green_2,
                                   command=self.get_excel_output)
        print_result.place(relx=0.33, rely=0.935, anchor="center")
        
        # دکمه خروج
        back_to_dashboard_btn = ctk.CTkButton(main_box, text="بازگشت", 
                                   width=120, height=50, 
                                   font=(None, 22, "bold"), 
                                   corner_radius=20, 
                                   text_color=colors.white, 
                                   bg_color=colors.white, 
                                   fg_color=colors.dark_green_5, 
                                   hover_color=colors.green_2,
                                   command=go_to_dashboard)
        back_to_dashboard_btn.place(relx=0.08, rely=0.935, anchor="center")

        def on_item_select(event):
            item_id = self.tree.focus()
            if item_id: open_bol()
        self.tree.bind("<Double-1>", on_item_select)

        # لوگو
        ck.make_image(main_box, asset_paths.Logo_path, 66, 68, 0.02, 0.04, "nw")
        
        #فیلد های گزارش‌گیری
        #شماره بارنامه
        self.bol_switch_bin = ctk.BooleanVar(value=False)
        self.bol_first_entry = make_entry_box(top_left_box, "شماره بارنامه", 30, 160, colors.dark_green_6, colors.light_gray_4, colors.dark_green_6, colors.black, 0.71, 0.8*0.245, "normal")
        self.bol_second_entry = make_entry_box(top_left_box, "", 30, 140, colors.dark_green_6, colors.light_gray_4, colors.dark_green_6, colors.black, None, None, "disabled")
        update_live_information(self.bol_first_entry)
        update_live_information(self.bol_second_entry)
        # update_live_information(self.bol_switch_bin)
        
        
        switch_key = ctk.CTkSwitch(top_left_box, button_color=colors.white, button_hover_color=colors.green_3,
                                    text="بازه", 
                                    text_color=colors.white,
                                    fg_color=colors.light_green_1,
                                    progress_color=colors.light_green_5,
                                    variable=self.bol_switch_bin, 
                                    command=lambda: self.toggle_mode_load_id(self.bol_first_entry, self.bol_second_entry, self.bol_switch_bin, 0.445, 0.8*0.245, "شماره بارنامه"))
        switch_key.place(relx=0.94, rely=0.8*0.245, anchor="center")

        #تاریخ ثبت
        self.register_date_bin = ctk.BooleanVar(value=False)
        self.register_date_btn = ctk.CTkButton(
            top_left_box,
            text="تاریخ ثبت",
            height=24,
            width=140,
            corner_radius=60,
            bg_color=colors.dark_green_6,
            border_color=colors.dark_green_6,
            fg_color=colors.light_gray_4,
            hover_color=colors.green_3,
            text_color=colors.gray_3,
            command=lambda:open_calendar(main_box, self.register_date_btn, 0.5, 0.5)
        )
        self.register_date_btn.place(relx=0.71, rely=0.32, anchor="center")
        update_live_information(self.register_date_btn)
        
        self.register_date_btn2 = ctk.CTkButton(
            top_left_box,
            text="تاریخ ثبت",
            height=24,
            width=140,
            corner_radius=60,
            state="disabled",
            bg_color=colors.dark_green_6,
            border_color=colors.dark_green_6,
            fg_color=colors.light_gray_4,
            hover_color=colors.green_3,
            text_color=colors.gray_3,
            command=lambda:open_calendar(main_box, self.register_date_btn2, 0.5, 0.5)
        )
        register_date_switch_key = ctk.CTkSwitch(top_left_box, button_color=colors.white, button_hover_color=colors.green_3,
                                    text="بازه", 
                                    text_color=colors.white,
                                    fg_color=colors.light_green_1,
                                    progress_color=colors.light_green_5,
                                    variable=self.register_date_bin, 
                                    command=lambda: self.toggle_mode(self.register_date_btn, self.register_date_btn2, register_date_switch_key, 0.445, 0.32, "تاریخ ثبت"))
        register_date_switch_key.place(relx=0.94, rely=0.32, anchor="center")
        
        #تاریخ صدور
        self.date_switch_bin = ctk.BooleanVar(value=False)
        self.sent_date_btn = ctk.CTkButton(
            top_left_box,
            text="تاریخ صدور",
            height=24,
            width=140,
            corner_radius=60,
            bg_color=colors.dark_green_6,
            border_color=colors.dark_green_6,
            fg_color=colors.light_gray_4,
            hover_color=colors.green_3,
            text_color=colors.gray_3,
            command=lambda:open_calendar(main_box, self.sent_date_btn, 0.5, 0.5)
        )
        self.sent_date_btn.place(relx=0.71, rely=0.41, anchor="center")
        update_live_information(self.sent_date_btn)
        
        self.sent_date_btn2 = ctk.CTkButton(
            top_left_box,
            text="تاریخ صدور",
            height=24,
            width=140,
            corner_radius=60,
            bg_color=colors.dark_green_6,
            border_color=colors.dark_green_6,
            fg_color=colors.light_gray_4,
            hover_color=colors.green_3,
            text_color=colors.gray_3,
            command=lambda:open_calendar(main_box, self.sent_date_btn2, 0.5, 0.5)
        )
        switch_key = ctk.CTkSwitch(top_left_box, button_color=colors.white, button_hover_color=colors.green_3,
                                    text="بازه", 
                                    text_color=colors.white,
                                    fg_color=colors.light_green_1,
                                    progress_color=colors.light_green_5,
                                    variable=self.date_switch_bin, 
                                    command=lambda: self.toggle_mode(self.sent_date_btn, self.sent_date_btn2, self.date_switch_bin, 0.445, 0.41, "تاریخ صدور"))
        switch_key.place(relx=0.94, rely=0.41, anchor="center")
        
        #تاریخ ارسال
        self.register_date_bin2 = ctk.BooleanVar(value=False)
        self.sent2_date_btn = ctk.CTkButton(
            top_left_box,
            text="تاریخ ارسال",
            height=24,
            width=140,
            corner_radius=60,
            bg_color=colors.dark_green_6,
            border_color=colors.dark_green_6,
            fg_color=colors.light_gray_4,
            hover_color=colors.green_3,
            text_color=colors.gray_3,
            command=lambda:open_calendar(main_box, self.sent2_date_btn, 0.5, 0.5)
        )
        self.sent2_date_btn.place(relx=0.71, rely=0.5, anchor="center")
        update_live_information(self.sent2_date_btn)
        
        self.sent2_date_btn2 = ctk.CTkButton(
            top_left_box,
            text="تاریخ ارسال",
            height=24,
            width=140,
            corner_radius=60,
            bg_color=colors.dark_green_6,
            border_color=colors.dark_green_6,
            fg_color=colors.light_gray_4,
            hover_color=colors.green_3,
            text_color=colors.gray_3,
            command=lambda:open_calendar(main_box, self.sent2_date_btn2, 0.5, 0.5)
        )
        register_date_switch_key2 = ctk.CTkSwitch(top_left_box, button_color=colors.white, button_hover_color=colors.green_3,
                                    text="بازه", 
                                    text_color=colors.white,
                                    fg_color=colors.light_green_1,
                                    progress_color=colors.light_green_5,
                                    variable=self.register_date_bin2, 
                                    command=lambda: self.toggle_mode(self.sent2_date_btn, self.sent2_date_btn2, self.register_date_bin2, 0.445, 0.5, "تاریخ ارسال"))
        register_date_switch_key2.place(relx=0.94, rely=0.5, anchor="center")
        
        #نام راننده
        self.driver_name_entry = make_entry_box(top_left_box, "نام راننده", 30, 145, colors.dark_green_6, colors.light_gray_4, colors.dark_green_6, colors.black, 0.15, 0.8*0.25, "normal")
        update_live_information(self.driver_name_entry)
        
        #کد ملی راننده
        self.driver_id_entry = make_entry_box(top_left_box, "کد ملی راننده", 30, 145, colors.dark_green_6, colors.light_gray_4, colors.dark_green_6, colors.black, 0.15, 0.8*0.38, "normal")
        update_live_information(self.driver_id_entry)
        
        #نام مالک
        self.owner_name_entry = make_entry_box(top_left_box, "نام مالک خودرو", 30, 145, colors.dark_green_6, colors.light_gray_4, colors.dark_green_6, colors.black, 0.15, 0.8*0.51, "normal")
        update_live_information(self.owner_name_entry)
        
        #کارت هوشمند راننده
        self.driver_smart_id_entry = make_entry_box(top_left_box, "کارت هوشمند راننده", 30, 145, colors.dark_green_6, colors.light_gray_4, colors.dark_green_6, colors.black, 0.15, 0.8*0.64, "normal")
        update_live_information(self.driver_smart_id_entry)
        
        #کارت هوشمند نفتکش
        self.car_smart_id_entry = make_entry_box(top_left_box, "کارت هوشمند نفتکش", 30, 145, colors.dark_green_6, colors.light_gray_4, colors.dark_green_6, colors.black, 0.15, 0.8*0.77, "normal")
        update_live_information(self.car_smart_id_entry)
        
        #شماره نفتکش
        self.licence_safety_end = make_entry_box(top_left_box, "شماره نفتکش", 30, 145, colors.dark_green_6, colors.light_gray_4, colors.dark_green_6, colors.black, 0.15, 0.8*0.9, "normal")
        update_live_information(self.licence_safety_end)
        
        def run_update(selected_value):
            self.update_table_with__new_data()
            
        def update_second_list_options(selected_value):
            sub_array = [item["sub_array"] for item in load_main_type if item["name"] == selected_value][0]
            self.load_name.configure(values = sub_array, state = "active")
            self.update_table_with__new_data()
            
        #کوره/رنگی
        main_type_options = [item["name"] for item in load_main_type]
        self.main_type_list = ctk.CTkOptionMenu(
            top_left_box,
            variable=ctk.StringVar(value="عنوان کالا"),
            values=main_type_options,
            width=165,
            height=27,
            anchor="e",
            fg_color=colors.light_gray_4,
            text_color=colors.black,
            button_color=colors.green_2,
            corner_radius=60,
            button_hover_color=colors.dark_green_3,
            command=update_second_list_options
        )
        self.main_type_list.place(relx=0.84, rely=0.8*0.085, anchor="center")
        
        #نام فرآورده
        self.load_name = ctk.CTkOptionMenu(
            top_left_box, 
            variable=ctk.StringVar(value="فرآورده"),
            width=165,
            height=27,
            state="disabled",
            anchor="e",
            fg_color=colors.light_gray_4,
            text_color=colors.black,
            button_color=colors.green_2,
            corner_radius=60,
            button_hover_color=colors.dark_green_3,
            command=run_update
        )
        self.load_name.place(relx=0.54, rely=0.8*0.085, anchor="center")
        
         #نوع ارسال
        sent_type_options = types
        sent_type_str = ctk.StringVar(value="نوع ارسال")
        self.sent_type = ctk.CTkOptionMenu(
            top_left_box, 
            width=200,
            variable=sent_type_str,
            values=sent_type_options,
            height=27,
            anchor="e",
            fg_color=colors.light_gray_4,
            text_color=colors.black,
            button_color=colors.green_2,
            corner_radius=60,
            button_hover_color=colors.dark_green_3,
            command=run_update
        )
        self.sent_type.place(relx=0.2, rely=0.8*0.085, anchor="center")
        
        #مقدار کرایه پرداختی/دریافتی        
        #توابع
        def on_cash_slider_change(value):
            self.cash_slider1_label.configure(text=cc.eng_to_persian_date(self.cash_format(f"{int(value)}")))
            if self.cash_slider2.get() < value:
                self.cash_slider2_label.configure(text=cc.eng_to_persian_date(self.cash_format(f"{int(value)}")))
                self.cash_slider2.set(value)
            self.update_table_with__new_data()
        def on_cash_slider_change2(value):
            self.cash_slider2_label.configure(text=cc.eng_to_persian_date(self.cash_format(f"{int(max(int(value), self.cash_slider1.get()))}")))
            if self.cash_slider2.get() < self.cash_slider1.get():
                self.cash_slider2_label.configure(cc.eng_to_persian_date(self.cash_format(str(self.cash_slider1.get()))))
                self.cash_slider2.set(self.cash_slider1.get())
            else:
                self.cash_slider2_label.configure(text=cc.eng_to_persian_date(self.cash_format(f"{int(value)}")))
            self.update_table_with__new_data()
        
        #اسلایدر اول
        self.cash_slider1 = ctk.CTkSlider(top_left_box, from_=self.find_endpoints_value("input_money")[0], to=self.find_endpoints_value("input_money")[1], 
                                     number_of_steps=200, width=170, 
                                     fg_color=colors.white,
                                     button_color=colors.light_green_5,
                                     progress_color=colors.green_5,
                                     button_hover_color=colors.green_2,
                                     command=on_cash_slider_change)
        self.cash_slider1.set(self.find_endpoints_value("input_money")[0])
        self.cash_slider1.place(relx=0.6, rely=0.6, anchor="center")
        update_live_information(self.cash_slider1)
        
        #اسلایدر دوم
        self.cash_slider2 = ctk.CTkSlider(top_left_box, from_=self.find_endpoints_value("input_money")[0], to=self.find_endpoints_value("input_money")[1], 
                                     number_of_steps=200, width=170, 
                                     fg_color=colors.white,
                                     button_color=colors.light_green_5,
                                     progress_color=colors.green_5,
                                     button_hover_color=colors.green_2,
                                     command=on_cash_slider_change2)
        self.cash_slider2.set(self.find_endpoints_value("input_money")[1])
        
        #عنوان کرایه(لیبل)
        cash_label = ctk.CTkLabel(top_left_box, text="کرایه", font=(None, 15), text_color=colors.white, bg_color=colors.dark_green_6)
        cash_label.place(relx=0.79, rely=0.6, anchor="center")
        
        #لیبل اسلایدر دوم
        cash_label2 = ctk.CTkLabel(top_left_box, text="تا کرایه", font=(None, 15), text_color=colors.white, bg_color=colors.dark_green_6)
        
        #لیبل اسلایدر اول
        self.cash_slider1_label = ctk.CTkLabel(top_left_box, 
                                               text_color=colors.white,
                                  text=cc.eng_to_persian_date(self.cash_format(f"{int(self.cash_slider1.get())}")),
                                  font=(None, 15))
        self.cash_slider1_label.place(relx=0.369, rely=0.6, anchor="center")
        
        #لیبل اسلایدر دوم
        self.cash_slider2_label = ctk.CTkLabel(top_left_box, 
                                               text_color=colors.white,
                                  text=cc.eng_to_persian_date(self.cash_format(f"{int(self.cash_slider1.get())}")),
                                  font=(None, 15))

        #سوییچ انتخاب بازه ای
        self.cash_switch_bin = ctk.BooleanVar(value=False)
        cash_switch = ctk.CTkSwitch(top_left_box, button_color=colors.white, button_hover_color=colors.green_3,
                                    text="بازه", 
                                    text_color=colors.white,
                                    fg_color=colors.light_green_1,
                                    progress_color=colors.light_green_5,
                                    variable=self.cash_switch_bin,
                                    command=lambda: self.toggle_mode_cash_slider(self.cash_slider2, 
                                                                                 self.cash_slider2_label, 
                                                                                 cash_label,
                                                                                 cash_label2, 
                                                                                 self.cash_switch_bin, 
                                                                                 0.665,
                                                                                 "کرایه"))
        cash_switch.place(relx=0.935, rely=0.6, anchor="center")
        
        #سوییچ انتخاب دریافتی/پرداختی
        self.input_output_bin = ctk.BooleanVar(value=False)
        input_output_switch = ctk.CTkSwitch(top_left_box, button_color=colors.white, button_hover_color=colors.green_3,
                                    text="دریافتی", 
                                    text_color=colors.white,
                                    fg_color=colors.light_green_1,
                                    progress_color=colors.light_green_5,
                                    variable=self.input_output_bin,
                                    command=lambda:self.toggle_mode_input_output(input_output_switch, self.input_output_bin))
        input_output_switch.place(relx=0.935, rely=0.665, anchor="center")

        #مبدأ
        self.start_route_entry = make_entry_box(top_left_box, "مبدأ", 30, 190, colors.dark_green_6, colors.light_gray_4, colors.dark_green_6, colors.black, 0.82, 0.77, "normal")
        update_live_information(self.start_route_entry)
        
        #مقصد
        self.end_route_entry = make_entry_box(top_left_box, "مقصد", 30, 190, colors.dark_green_6, colors.light_gray_4, colors.dark_green_6, colors.black, 0.48, 0.77, "normal")
        update_live_information(self.end_route_entry)
        
        #مقدار محموله واقعی/طبیعی     
        #توابع
        def on_load_slider_change(value):
            self.load_slider1_label.configure(text=cc.eng_to_persian_date(self.cash_format(f"{int(value)}")))
            if self.load_slider2.get() < value:
                self.load_slider2_label.configure(text=cc.eng_to_persian_date(self.cash_format(f"{int(value)}")))
                self.load_slider2.set(value)
            self.update_table_with__new_data()
            
        def on_load_slider_change2(value):
            self.load_slider2_label.configure(text=cc.eng_to_persian_date(self.cash_format(f"{int(max(int(value), self.load_slider1.get()))}")))
            if self.load_slider2.get() < self.load_slider1.get():
                self.load_slider2_label.configure(cc.eng_to_persian_date(self.cash_format(str(self.load_slider1.get()))))
                self.load_slider2.set(self.load_slider1.get())
            else:
                self.load_slider2_label.configure(text=cc.eng_to_persian_date(self.cash_format(f"{int(value)}")))
            self.update_table_with__new_data()
        
        #اسلایدر اول
        self.load_slider1 = ctk.CTkSlider(top_left_box, from_=self.find_endpoints_value("true_value")[0], to=self.find_endpoints_value("true_value")[1], 
                                     number_of_steps=200, width=170, 
                                     fg_color=colors.white,
                                     button_color=colors.light_green_5,
                                     progress_color=colors.green_5,
                                     button_hover_color=colors.green_2,
                                     command=on_load_slider_change)
        self.load_slider1.set(self.find_endpoints_value("true_value")[0])
        self.load_slider1.place(relx=0.55, rely=0.8856, anchor="center")
        
        #اسلایدر دوم
        self.load_slider2 = ctk.CTkSlider(top_left_box, from_=self.find_endpoints_value("true_value")[0], to=self.find_endpoints_value("true_value")[1], 
                                     number_of_steps=200, width=170, 
                                     fg_color=colors.white,
                                     button_color=colors.light_green_5,
                                     progress_color=colors.green_5,
                                     button_hover_color=colors.green_2,
                                     command=on_load_slider_change2)
        self.load_slider2.set(self.find_endpoints_value("true_value")[1])
        
        #عنوان محموله(لیبل)
        load_label = ctk.CTkLabel(top_left_box, text="مقدار", font=(None, 15), text_color=colors.white, bg_color=colors.dark_green_6)
        load_label.place(relx=0.745, rely=0.8856, anchor="center")
        
        #لیبل اسلایدر دوم
        load_label2 = ctk.CTkLabel(top_left_box, text="تا مقدار", font=(None, 15), text_color=colors.white, bg_color=colors.dark_green_6)
        
        #لیبل اسلایدر اول
        self.load_slider1_label = ctk.CTkLabel(top_left_box, 
                                               text_color=colors.white,
                                  text=cc.eng_to_persian_date(self.cash_format(f"{int(self.load_slider1.get())}")),
                                  font=(None, 15))
        self.load_slider1_label.place(relx=0.35, rely=0.8856, anchor="center")
        
        #لیبل اسلایدر دوم
        self.load_slider2_label = ctk.CTkLabel(top_left_box, 
                                    text_color=colors.white,
                                  text=cc.eng_to_persian_date(self.cash_format(f"{int(self.load_slider2.get())}")),
                                  font=(None, 15))

        #سوییچ انتخاب بازه ای
        self.load_switch_bin = ctk.BooleanVar(value=False)
        load_switch = ctk.CTkSwitch(top_left_box, button_color=colors.white, button_hover_color=colors.green_3,
                                    text="بازه", 
                                    text_color=colors.white,
                                    fg_color=colors.light_green_1,
                                    progress_color=colors.light_green_5,
                                    variable=self.load_switch_bin,
                                    command=lambda: self.toggle_mode_load_slider(self.load_slider2, 
                                                                                 self.load_slider2_label, 
                                                                                 load_label,
                                                                                 load_label2, 
                                                                                 self.load_switch_bin, 
                                                                                 0.9506,
                                                                                 "مقدار"))
        load_switch.place(relx=0.905, rely=0.95, anchor="center")
        
        #سوییچ انتخاب حمل‌شده/واقعی
        self.real_gone_bin = ctk.BooleanVar(value=False)
        real_gone_switch = ctk.CTkSwitch(top_left_box, button_color=colors.white, button_hover_color=colors.green_3,
                                    text="واقعی", 
                                    text_color=colors.white,
                                    fg_color=colors.light_green_1,
                                    progress_color=colors.light_green_5,
                                    variable=self.real_gone_bin,
                                    command=lambda:self.toggle_mode_real_gone(real_gone_switch, self.real_gone_bin))
        real_gone_switch.place(relx=0.905, rely=0.82*1.08, anchor="center")

        
        # دکمه ریست
        reset_btn = ctk.CTkButton(top_left_box, text="دوباره", 
                                   width=145, height=35, 
                                   font=(None, 15, "bold"), 
                                   corner_radius=20, 
                                   text_color=colors.white, 
                                   bg_color=colors.dark_green_6, 
                                   fg_color=colors.green_2, 
                                   hover_color=colors.light_green_5,
                                   command=self.reset_page)
        reset_btn.place(relx=0.15, rely=0.9, anchor="center")
                
        #نوع بارنامه ها
        self.load_type = ck.make_list(main_box, "بارنامه‌های فعال", ["بارنامه‌های فعال", "بارنامه‌های آرشیو", "همه‌ بارنامه‌ها"], 150, 30, colors.black,
                     colors.white, colors.light_green_2, colors.green_4, colors.green_1, 20, 
                     self.update_bol_box, "e", 0.985, 0.14, "e")
        
        def get_default_deal():
            if(len(current_deals) > 0): 
                return 'همه قراردادها', [(deal["name"] + "/" + cc.eng_to_persian_date(deal["id"])) for deal in current_deals] + ['همه قراردادها']
            else: return "عدم وجود قرارداد", []
            
        #انتخاب قرارداد
        deafult_deal, total_deals = get_default_deal()
        self.deal = ck.make_list(main_box, deafult_deal, total_deals, 150, 30, colors.black,
                     colors.white, colors.light_green_2, colors.green_4, colors.green_1, 20, 
                     self.update_bol_box, "e", 0.65, 0.14, "e")
        #انتخاب نماینده
        self.agent = ck.make_list(main_box, "همه نماینده‌ها", agents + ['همه نماینده‌ها'], 150, 30, colors.black,
                     colors.white, colors.light_green_2, colors.green_4, colors.green_1, 20, 
                     self.update_bol_box, "e", 0.818, 0.14, "e")
        
        # دکمه بستن صورت و بایگانی
        make_group_btn = ck.make_button(main_box, "صورت کردن", 150, 50, (None, 22, "bold"), 20, 
                       colors.white, colors.white, colors.dark_green_6, colors.green_4, self.make_group, 0.445, 0.935, "center")
        make_group_type2_btn = ck.make_button(main_box, "صورت دستی", 150, 50, (None, 22, "bold"), 20, 
                       colors.white, colors.white, colors.dark_green_6, colors.green_4, self.make_group_type2, 0.275, 0.935, "e")
        
        if(self.deal.get() == 'عدم وجود قرارداد'):
            make_group_btn.configure(state='disabled')
        #خط عمودی باکس بالا سمت چپ
        circle_canvas = ctk.CTkCanvas(top_left_box, height=float(top_left_box.winfo_screenwidth() / 3.35), width=2, bg=colors.white, highlightthickness=0)
        circle_canvas.place(relx=0.29, rely=0.135)
        
        #خط افتی بالای اسلایدر کرایه ها
        circle_canvas = ctk.CTkCanvas(top_left_box, width=float(top_left_box.winfo_screenwidth()), height=2, bg=colors.white, highlightthickness=0)
        circle_canvas.place(relx=0, rely=0.8*0.165)
                
        #خط افتی بالای اسلایدر کرایه ها
        circle_canvas = ctk.CTkCanvas(top_left_box, width=float(top_left_box.winfo_screenwidth()), height=2, bg=colors.white, highlightthickness=0)
        circle_canvas.place(relx=0.29, rely=0.55)

        #خط افتی بالای مبدا و مقصد
        circle_canvas = ctk.CTkCanvas(top_left_box, width=float(top_left_box.winfo_screenwidth()), height=2, bg=colors.white, highlightthickness=0)
        circle_canvas.place(relx=0.29, rely=0.703)
        
        #خط افتی بالای مقدار محموله
        circle_canvas = ctk.CTkCanvas(top_left_box, width=float(top_left_box.winfo_screenwidth()), height=2, bg=colors.white, highlightthickness=0)
        circle_canvas.place(relx=0.29, rely=0.83)
                
        #خط افتی بالای تاریخ ها
        circle_canvas = ctk.CTkCanvas(top_left_box, width=float(top_left_box.winfo_screenwidth()), height=2, bg=colors.white, highlightthickness=0)
        circle_canvas.place(relx=0.29, rely=0.26)
        
        #لیبل های باکس پایین چپ
        #جمع وزن(لیتر)
        self.sum_weights_litr = ctk.CTkLabel(buttom_left_box, text=f"جمع وزن حمل شده(لیتر): {self.cash_format(self.eng_to_persian_num(self.find_initial_result("load_weight")))}", font=(None, 12,"bold"), 
                                   text_color=colors.black, 
                                   bg_color=colors.light_green_4)
        self.sum_weights_litr.place(relx=0.97, rely=0.25, anchor="e")
        
        #جمع وزن(کیلو)
        self.sum_weights_kilo = ctk.CTkLabel(buttom_left_box, text=f"جمع وزن حمل شده(کیلو): {self.cash_format(self.eng_to_persian_num(self.find_initial_result("load_weight")))}", font=(None, 12,"bold"), 
                                   text_color=colors.black, 
                                   bg_color=colors.light_green_4)
        self.sum_weights_kilo.place(relx=0.97, rely=0.75, anchor="e")
        
        #جمع وزن(لیتر)
        self.sum_weights_litr2 = ctk.CTkLabel(buttom_left_box, text=f"جمع وزن واقعی(لیتر): {self.cash_format(self.eng_to_persian_num(self.find_initial_result("true_value")))}", font=(None, 12,"bold"), 
                                   text_color=colors.black, 
                                   bg_color=colors.light_green_4)
        self.sum_weights_litr2.place(relx=0.62, rely=0.25, anchor="e")
        
        #جمع وزن(کیلو)
        self.sum_weights_kilo2 = ctk.CTkLabel(buttom_left_box, text=f"جمع وزن واقعی(کیلو): {self.cash_format(self.eng_to_persian_num(self.find_initial_result("true_value")))}", font=(None, 12,"bold"), 
                                   text_color=colors.black, 
                                   bg_color=colors.light_green_4)
        self.sum_weights_kilo2.place(relx=0.62, rely=0.75, anchor="e")
        
        #جمع کرایه دریافتی
        self.sum_cash_input = ctk.CTkLabel(buttom_left_box, text=f"جمع کرایه دریافتی: {self.cash_format(self.eng_to_persian_num(self.find_initial_result("input_money")))}", font=(None, 12,"bold"), 
                                   text_color=colors.black, 
                                   bg_color=colors.light_green_4)
        self.sum_cash_input.place(relx=0.32, rely=0.25, anchor="e")
        
        #جمع کرایه پرداختی
        self.sum_cash_output = ctk.CTkLabel(buttom_left_box, text=f"جمع کرایه پرداختی: {self.cash_format(self.eng_to_persian_num(self.find_initial_result("payed_money")))}", font=(None, 12,"bold"), 
                                   text_color=colors.black, 
                                   bg_color=colors.light_green_4)
        self.sum_cash_output.place(relx=0.32, rely=0.75, anchor="e")
        
